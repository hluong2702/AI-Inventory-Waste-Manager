import openpyxl

path = "/Users/hoangluong/Downloads/Template-Technical.xlsx"
workbook = openpyxl.load_workbook(path, data_only=False)
for sheet in workbook.worksheets:
    print("###", sheet.title)
    for row in sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row, 150),
        max_col=min(sheet.max_column, 80),
    ):
        values = [
            f"{cell.coordinate}={str(cell.value)[:250]}"
            for cell in row
            if cell.value not in (None, "")
        ]
        if values:
            print(" | ".join(values))
